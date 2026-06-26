"""Read and join the three NTT source workbooks into `Village` records.

Inputs (English names in docs/ntt_data_integration.md):
  - Data potensi desa.xlsx            -> households / electrification + economic sector
  - NTT_Data_Desa.xlsx (GHI_per_Desa) -> per-village GHI + lat/long
  - (the KDKMP calculator is consumed by tools.ntt.calculators, not here)

Villages are joined on normalised (Kabupaten, Kecamatan, Desa), filtered to the
four Timor kabupaten, and tagged with an archetype. Unmatched GHI falls back to
the kabupaten mean (then the NTT mean), flagged on the record.
"""

from __future__ import annotations

import re
from collections import defaultdict

import openpyxl

from .archetypes import assign_archetype, DEFAULT_GHI
from .calculators.base import Village

# The four Timor kabupaten in scope (normalised).
TIMOR_KABUPATEN = {
    "KUPANG",
    "TIMOR TENGAH SELATAN",
    "TIMOR TENGAH UTARA",
    "BELU",
}


def norm(s) -> str:
    """Normalise a name for joining: upper-case, collapse whitespace, strip."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().upper())


def _num(x) -> int:
    try:
        return int(x) if x is not None else 0
    except (ValueError, TypeError):
        return 0


def load_potensi(path: str, kabupaten: set[str] | None = None) -> dict:
    """Return {(kab,kec,desa): record} from Data potensi desa.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet2"]
    out = {}
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or r[0] is None:
            continue
        kab = norm(r[1])
        if kabupaten and kab not in kabupaten:
            continue
        kec, desa = norm(r[2]), norm(r[3])
        out[(kab, kec, desa)] = {
            "kab": kab, "kec": kec, "desa": desa,
            "sector": r[5], "subsector": r[7], "commodity": r[8],
            "hh_pln": _num(r[9]), "hh_nonpln": _num(r[10]), "hh_unelec": _num(r[11]),
        }
    wb.close()
    return out


def load_ghi(path: str, kabupaten: set[str] | None = None) -> tuple[dict, dict]:
    """Return ({(kab,kec,desa): {ghi,lat,lon}}, {kab: mean_ghi}) from NTT_Data_Desa.xlsx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["GHI_per_Desa"]
    out, by_kab = {}, defaultdict(list)
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        kab = norm(r[4])
        if kabupaten and kab not in kabupaten:
            continue
        kec, desa = norm(r[2]), norm(r[1])
        ghi = r[8]
        out[(kab, kec, desa)] = {"ghi": ghi, "lon": r[6], "lat": r[7]}
        if ghi is not None:
            by_kab[kab].append(ghi)
    wb.close()
    kab_mean = {k: sum(v) / len(v) for k, v in by_kab.items() if v}
    return out, kab_mean


def load_timor_villages(potensi_path: str, ghi_path: str,
                        kabupaten: set[str] | None = None) -> tuple[list[Village], dict]:
    """Join the two workbooks into Village records for the Timor kabupaten.

    Returns (villages, stats). `stats` reports match rates for logging."""
    kabupaten = kabupaten or TIMOR_KABUPATEN
    potensi = load_potensi(potensi_path, kabupaten)
    ghi, kab_mean = load_ghi(ghi_path, kabupaten)

    # secondary index for (kab, desa) fallback when kecamatan spelling differs
    ghi_by_kab_desa = {}
    for (kab, kec, desa), g in ghi.items():
        ghi_by_kab_desa.setdefault((kab, desa), g)

    villages, matched, fell_back = [], 0, 0
    for key, p in sorted(potensi.items()):
        kab, kec, desa = key
        g = ghi.get(key) or ghi_by_kab_desa.get((kab, desa))
        if g and g.get("ghi") is not None:
            gv, lat, lon, is_match = g["ghi"], g.get("lat"), g.get("lon"), True
            matched += 1
        else:
            gv = kab_mean.get(kab, DEFAULT_GHI)
            lat = lon = None
            is_match = False
            fell_back += 1
        # Classify from the specific subsector/commodity. The umbrella sector
        # ("Pertanian, kehutanan, dan perikanan") contains "perikanan"/"kehutanan"
        # and must NOT drive the archetype — pass it only for non-agriculture rows.
        is_agri = str(p["sector"] or "").strip().lower().startswith("pertanian")
        texts = [p["subsector"], p["commodity"]]
        if not is_agri:
            texts.append(p["sector"])
        arch = assign_archetype(*texts)
        # Agriculture village with no usable subsector/commodity -> regional
        # default is food crops (maize/rice dominate Timor).
        if arch.name == "default" and is_agri:
            from .archetypes import REGISTRY
            arch = REGISTRY["rice"]
        villages.append(Village(
            kabupaten=kab, kecamatan=kec, desa=desa,
            households=p["hh_pln"] + p["hh_nonpln"] + p["hh_unelec"],
            hh_pln=p["hh_pln"], hh_nonpln=p["hh_nonpln"], hh_unelectrified=p["hh_unelec"],
            ghi=round(gv, 4), lat=lat, lon=lon, archetype=arch.name, matched_ghi=is_match,
        ))
    stats = {"villages": len(villages), "ghi_matched": matched,
             "ghi_fell_back": fell_back, "kabupaten": sorted(kabupaten)}
    return villages, stats
